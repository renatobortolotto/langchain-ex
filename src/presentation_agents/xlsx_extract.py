from __future__ import annotations

import json
import re
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple, Union

try:  # pragma: no cover
    from openpyxl import load_workbook as _openpyxl_load_workbook  # type: ignore
    from openpyxl.utils.cell import range_boundaries as _openpyxl_range_boundaries  # type: ignore
    from openpyxl.utils.exceptions import InvalidFileException  # type: ignore
except Exception:  # pragma: no cover
    _openpyxl_load_workbook = None
    _openpyxl_range_boundaries = None
    InvalidFileException = None


@dataclass(frozen=True)
class ExtractSpec:
    id: str
    labels_range: str
    values_range: str
    sheet: Optional[str] = None


_A1_CELL_RE = re.compile(r"^\$?([A-Za-z]+)\$?(\d+)$")


def _col_letters_to_index(letters: str) -> int:
    letters = letters.strip().upper()
    if not letters.isalpha():
        raise ValueError(f"Coluna invalida: {letters!r}")
    col = 0
    for ch in letters:
        col = col * 26 + (ord(ch) - ord("A") + 1)
    return col


def _a1_to_rowcol(a1: str) -> Tuple[int, int]:
    a1 = a1.strip()
    m = _A1_CELL_RE.match(a1)
    if not m:
        raise ValueError(f"Celula A1 invalida: {a1!r}")
    col_letters, row_s = m.groups()
    return int(row_s), _col_letters_to_index(col_letters)


def _range_boundaries(a1_range: str) -> Tuple[int, int, int, int]:
    a1_range = a1_range.strip()
    if not a1_range:
        raise ValueError("Range vazio")

    if _openpyxl_range_boundaries is not None:  # pragma: no cover
        min_col, min_row, max_col, max_row = _openpyxl_range_boundaries(a1_range)
        return int(min_col), int(min_row), int(max_col), int(max_row)

    if ":" not in a1_range:
        row, col = _a1_to_rowcol(a1_range)
        return col, row, col, row

    left, right = [p.strip() for p in a1_range.split(":", 1)]
    row1, col1 = _a1_to_rowcol(left)
    row2, col2 = _a1_to_rowcol(right)
    min_col, max_col = sorted((col1, col2))
    min_row, max_row = sorted((row1, row2))
    return min_col, min_row, max_col, max_row


def _load_workbook(*, filename, data_only: bool = True):
    if _openpyxl_load_workbook is None:  # pragma: no cover
        raise RuntimeError("Dependencia 'openpyxl' nao instalada")

    import zipfile

    try:
        return _openpyxl_load_workbook(filename=filename, data_only=data_only)
    except (zipfile.BadZipFile, OSError, ValueError) as exc:
        raise ValueError("Arquivo enviado nao e um XLSX valido") from exc
    except Exception as exc:
        if InvalidFileException is not None and isinstance(exc, InvalidFileException):  # pragma: no cover
            raise ValueError("Arquivo enviado nao e um XLSX valido") from exc
        raise ValueError("Arquivo enviado nao e um XLSX valido") from exc


def _read_range_2d(ws, a1_range: str) -> List[List[Any]]:
    a1_range = a1_range.strip()
    if not a1_range:
        raise ValueError("Range vazio")

    min_col, min_row, max_col, max_row = _range_boundaries(a1_range)
    out: List[List[Any]] = []
    for r in range(min_row, max_row + 1):
        row_vals: List[Any] = []
        for c in range(min_col, max_col + 1):
            row_vals.append(ws.cell(row=r, column=c).value)
        out.append(row_vals)
    return out


def _to_1d(values_2d: List[List[Any]]) -> List[Any]:
    if not values_2d:
        return []
    rows = len(values_2d)
    cols = len(values_2d[0]) if rows else 0

    if rows == 1:
        return list(values_2d[0])
    if cols == 1:
        return [r[0] for r in values_2d]

    flat: List[Any] = []
    for row in values_2d:
        flat.extend(row)
    return flat


def _coerce_labels(values: Sequence[Any]) -> List[str]:
    return ["" if value is None else str(value) for value in values]


def _coerce_values(values: Sequence[Any], *, strict: bool) -> List[Optional[float]]:
    out: List[Optional[float]] = []
    for value in values:
        if value is None or (isinstance(value, str) and value.strip() == ""):
            out.append(None)
            continue
        if isinstance(value, (int, float)):
            out.append(float(value))
            continue
        try:
            out.append(float(value))
        except (TypeError, ValueError):
            if strict:
                raise ValueError(f"Valor nao numerico: {value!r}")
            out.append(None)
    return out


def extract_workbook_to_dict(
    wb,
    specs: Sequence[ExtractSpec],
    *,
    default_sheet: Optional[str] = None,
    strict_numbers: bool = False,
    include_meta: bool = False,
    lowercase_fields: bool = False,
) -> Dict[str, Any]:
    result: Dict[str, Any] = {}

    labels_key = "labels" if lowercase_fields else "Labels"
    values_key = "values" if lowercase_fields else "Values"
    sheet_key = "sheet" if lowercase_fields else "Sheet"
    ranges_key = "ranges" if lowercase_fields else "Ranges"

    for spec in specs:
        sheet_name = spec.sheet or default_sheet
        if not sheet_name:
            raise ValueError(
                f"Spec {spec.id!r} nao tem sheet e nenhum default_sheet foi informado"
            )
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Aba nao encontrada: {sheet_name!r} (spec={spec.id!r}). Disponiveis: {wb.sheetnames}"
            )

        ws = wb[sheet_name]
        labels_raw = _to_1d(_read_range_2d(ws, spec.labels_range))
        values_raw = _to_1d(_read_range_2d(ws, spec.values_range))

        payload: Dict[str, Any] = {
            labels_key: _coerce_labels(labels_raw),
            values_key: _coerce_values(values_raw, strict=strict_numbers),
        }
        if include_meta:
            payload[sheet_key] = sheet_name
            payload[ranges_key] = {
                labels_key: spec.labels_range,
                values_key: spec.values_range,
            }

        result[spec.id] = payload

    return result


def extract_xlsx_bytes_to_dict(
    xlsx_bytes: bytes,
    specs: Sequence[ExtractSpec],
    *,
    default_sheet: Optional[str] = None,
    strict_numbers: bool = False,
    include_meta: bool = False,
    lowercase_fields: bool = False,
) -> Dict[str, Any]:
    if not xlsx_bytes:
        raise ValueError("XLSX vazio")

    wb = _load_workbook(filename=BytesIO(xlsx_bytes), data_only=True)
    return extract_workbook_to_dict(
        wb,
        specs,
        default_sheet=default_sheet,
        strict_numbers=strict_numbers,
        include_meta=include_meta,
        lowercase_fields=lowercase_fields,
    )


def parse_specs_json(path: Union[str, Path]) -> List[ExtractSpec]:
    path = Path(path)
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        raise ValueError("O arquivo de specs deve ser uma lista JSON")
    return parse_specs_data(data)


def parse_specs_data(data: Iterable[object]) -> List[ExtractSpec]:
    out: List[ExtractSpec] = []
    for item in data:
        if not isinstance(item, dict):
            raise ValueError("Cada spec deve ser um objeto JSON")

        spec_id = str(item.get("id") or item.get("ID") or "").strip()
        if not spec_id:
            raise ValueError("Spec sem 'id'")

        labels_range = item.get("labels_range") or item.get("labels")
        values_range = item.get("values_range") or item.get("values")
        sheet = item.get("sheet")
        if not labels_range or not values_range:
            raise ValueError(
                f"Spec {spec_id!r} precisa de labels_range e values_range (ou labels/values)"
            )

        out.append(
            ExtractSpec(
                id=spec_id,
                labels_range=str(labels_range),
                values_range=str(values_range),
                sheet=str(sheet) if sheet else None,
            )
        )
    return out
